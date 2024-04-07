#Nexttodo: Fix GetMix first with other jobboards via print board funtion
#Todo: refractor code for numpy for loop
# Implement Solid Principles

import webbrowser
from openpyxl import load_workbook
import sys, os
from numpy import random 
from googlesearch import search
#import markdownify

data_file =  os.path.abspath(os.getcwd())  + "\\Job Search Spring 2023.xlsx"
wb = load_workbook(data_file)
Company_ws = wb['Companies']
JobBoard_ws = wb["JobBoard"]

def isBlank(ws) -> bool:
   rows = list(ws)
   empty = {}
   
   for cells in rows:
      if not cells[0].value or not cells[1].value:
         empty[cells[0].value] = cells[1].value
   if len(empty) > 0:
      print(empty)
      return True
   else:
      return False
def deleteBlank(ws) -> None:
   for row in ws.iter_rows():
      if not all(cell.value for cell in row):
         ws.delete_row(row[0].row, 1)
         deleteBlank(ws)
 
def deleteRepeats() -> bool:
   #In case multiple worksheet are available
   #for ws in wb:
    # print(ws)
   #sheet = input("Pick a sheet")
   #ws = wb[sheet]
   ws = wb["Companies"]
   all_rows = list(ws.rows)
   repeat = {}
   for cells in range(0,len(all_rows)):
      for next in range(cells+1,len(all_rows)):
         if str(all_rows[cells][0].value).upper() == str(all_rows[next][0].value).upper():
            repeat[next] = all_rows[next][0].value
            ws.delete_rows(next, 1)
   if len(repeat) > 0:
      #To Fix
      #prompt =  "Following Rows were deleted: " + str(list(repeat.keys)) + str(list(repeat.values))
      #print(prompt)
      return True
   else:
      return False
# Retriving Functions
# Retriving name of company
def GetCompany(size = -1):
   company_rows = list(Company_ws.rows)
   return [cells[0].value for cells in company_rows[1:size]]
#Retriving name of company and thier career sites
def GetCareerSites(size = -1):   
   company_rows = list(Company_ws.rows)
   return {cells[0].value: "site:" + str(cells[1].value) + keyword for cells in company_rows[1:size]}
#Retriving name site of job boards
#ToDo: Change to specifics?  Gets all of it for now
def GetJobBoards(size = -1):
   Board_rows = list(JobBoard_ws)
   return {cells[0].value:cells[1].value for cells in Board_rows[1:size]}
# Creating a query for comapnies based on certain job boards
# ToDo: multiple jobboards? prob not
def GetMix():
   boards = GetJobBoards()
   
   prompt = "Pick one of these Job Boards:\n" + str(list(boards.keys())) + ": \n"
   board = input(prompt)

   company_rows = list(Company_ws.rows)

   prompt = "How many companies do you want to look through from " + str(len(company_rows)) + ": \n"
   sites_lengths = input(prompt)
   jobSites = ["site:" + boards[board] + keyword +" AND " + str(cells[0].value) for cells in company_rows[1:]]
      #word = " AND " + str(cells[0].value)
      #jobs.append("site:" + boards[board] + keyword + word)
   
   random.shuffle(jobSites)
   return jobSites[:int(sites_lengths)]

#Finding career sites from google of companies
def findCareerSite(sites, size = 10):
   print(size,"sites out of", len(sites))
   save = []
   for site in sites[:size]:
      query = str(site) + " careersites"
      for j in search(query, tld="co.in", num=3, stop=4, pause=2):
         print(j)
         save.append(j)
   return save

#Adding functions
def addCareerSite(sites) -> None:
   directory  = os.path.abspath(os.getcwd()) 
   NewSites = wb.create_sheet("New Company Sites")
   
   save = findCareerSite(sites, len(sites))
   saves = {i:save[i] for i in range(0,len(save))}
   #Todo: make it a column instead of row
   NewSites.append(saves)
   path = directory + "\\Test.xlsx"
   print("Saving New Doc to Test.xls")
   wb.save(path)
# def addJobs
# def addBoards

def PrintSites(sites) -> None:
   print("There are ", len(sites), " Websites")
   print("we would have opened these sites:")
   for site in sites:
      print(site)
def OpenSites(sites, size=10) -> None:
   #query = ""
   print("Amount of sites: ", len(sites))
      
   random.shuffle(sites)
   for site in sites[:int(size)]:
      webbrowser.open_new_tab(site)  # Go to example.com
      print(site)

# Googles size with flag being default, 
def GoogleSites(sites, size=10) -> None:
   sites = list(sites.keys())
   print("Googling...... ",size, " out of ", len(sites))
   for site in sites[:size]:
      for j in search(site, tld="co.in", num=1, stop=4, pause=2):
         print(j)

if __name__ == "__main__":
   keyword = " Cybersecurity AND RecentGrad"
   companies = []

   directory = os.path.abspath(os.getcwd()) 
   data_file = directory + "\\JobSearch.xlxs"
   if '-custom' in sys.argv:
      keyword = input("Enter keyword: ")
   if '-teal' in sys.argv:
      OpenSites("https://app.tealhq.com/home",1)
   if '-c' in sys.argv:
      print("we are Opening Sites")
      print("--------------------")
      c = GetCareerSites()
      OpenSites(list(c.values()))
   elif '-b' in sys.argv:
      print("Time to check Job Boards")
      print("--------------------")
      boards = GetJobBoards()
      OpenSites(list(boards.values()), len(boards))
   elif '-m' in sys.argv:
      print("we are mixing up")
      print("--------------------")
      m = GetMix()
      GoogleSites(m)
   elif '-d' in sys.argv:
      print("Starting deleting attempt")
      print("----------------------")
      deleteRepeats()
      path = "C:\\Users\\Tgang\\Downloads\\Test.xlsx"
      wb.save(path)
   elif '-s' in sys.argv:
      print("These are your Career Sites")
      print("----------------------------")
      m = addCareerSite(GetCareerSites() )
   
   elif '-h' in sys.argv:
      print('Job Search is a supplementary automation for JobSearch xml document (please update path)')
   
   elif '-t' in sys.argv:
      if not(isBlank(Company_ws)):
         c = GetCareerSites() 
         print("Testing Opening Career sites")
         print("There are ", len(c)," Websites and ", len(list(Company_ws.rows)), " Rows")
         print("we would have opened these sites:")
         for site in c:
            print(site)
      print('---------------------------')
      if not(isBlank(JobBoard_ws)):
         b = GetJobBoards()
         print("Testing Opening Career sites")
         print("There are ", len(b), " Websites and ", len(list(Company_ws.rows)), " Rows")
         print("we would have opened these sites:")
         for site in b:
            print(site)
      print('--------------------------------')
      m = GetMix()
      print("There are ", len(m)," Websites")
      print("we would have opened these sites:")
      for site in m:
         print(site)
   else:
      #gets a list of 10 career sites from current list as default
      companies = GetCompany()
      random.shuffle(companies)
      default = findCareerSite(companies,10)
      check = input("Open these sites? y/N: ")
      if check == ('y' or 'yes'):
         OpenSites(default)
   exit() 